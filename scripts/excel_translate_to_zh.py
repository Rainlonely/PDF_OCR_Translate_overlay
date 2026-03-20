#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import time
from collections import Counter
from pathlib import Path
from typing import Dict

from deep_translator import GoogleTranslator
from openpyxl import load_workbook


def load_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def looks_like_non_natural_text(text: str) -> bool:
    s = text.strip()
    if not s:
        return True
    if re.fullmatch(r"https?://\S+", s, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"[\w.+-]+@[\w.-]+\.\w+", s):
        return True
    if re.fullmatch(r"[A-Z0-9_.\-/]{6,}", s):
        return True
    return False


def should_translate_text(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = normalize_text(value)
    if not text:
        return False
    if looks_like_non_natural_text(text):
        return False
    if not re.search(r"[A-Za-z]", text):
        return False
    return True


def split_chunks(text: str, max_len: int = 1800) -> list[str]:
    paras = re.split(r"\n\s*\n", text)
    chunks: list[str] = []
    cur = ""
    for para in paras:
        para = para.strip()
        if not para:
            continue
        candidate = f"{cur}\n\n{para}".strip() if cur else para
        if len(candidate) <= max_len:
            cur = candidate
            continue
        if cur:
            chunks.append(cur)
        if len(para) <= max_len:
            cur = para
            continue
        lines = para.splitlines()
        buf = ""
        for line in lines:
            candidate = f"{buf}\n{line}".strip() if buf else line
            if len(candidate) <= 1200:
                buf = candidate
            else:
                if buf:
                    chunks.append(buf)
                buf = line
        cur = buf
    if cur:
        chunks.append(cur)
    return chunks


def clean_translation(text: str) -> str:
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            lines.append("")
            continue
        if s.count("?") / max(1, len(s)) >= 0.4:
            continue
        lines.append(line)
    return "\n".join(lines).strip() or text


def translate_text(text: str, translator: GoogleTranslator, cache: Dict[str, str]) -> str:
    text = normalize_text(text)
    if not text:
        return text
    if text in cache:
        return cache[text]

    translated_parts: list[str] = []
    for chunk in split_chunks(text):
        translated = None
        for _ in range(4):
            try:
                translated = translator.translate(chunk)
                break
            except Exception:
                time.sleep(1.0)
        if translated is None:
            translated = "[翻译失败]\n" + chunk
        translated_parts.append(translated)

    result = clean_translation("\n\n".join(translated_parts))
    cache[text] = result
    return result


def translate_missing_texts(
    texts: list[str],
    translator: GoogleTranslator,
    cache: Dict[str, str],
    cache_path: Path,
) -> None:
    pending = [normalize_text(text) for text in texts if normalize_text(text) and normalize_text(text) not in cache]
    if not pending:
        return

    batch_size = 15
    for idx in range(0, len(pending), batch_size):
        batch = pending[idx : idx + batch_size]
        translated_batch = None
        for _ in range(4):
            try:
                translated_batch = translator.translate_batch(batch)
                break
            except Exception:
                time.sleep(1.0)
        if not translated_batch or len(translated_batch) != len(batch):
            for text in batch:
                translate_text(text, translator, cache)
        else:
            for original, translated in zip(batch, translated_batch):
                cache[original] = clean_translation(translated)
        print(f"  batch translated: {min(idx + len(batch), len(pending))}/{len(pending)}", flush=True)
        save_json(cache_path, cache)


def translate_workbook(input_path: Path, output_path: Path, cache_path: Path) -> None:
    wb = load_workbook(
        filename=str(input_path),
        keep_vba=input_path.suffix.lower() == ".xlsm",
    )
    translator = GoogleTranslator(source="auto", target="zh-CN")
    cache: Dict[str, str] = load_json(cache_path, {})

    texts_to_translate: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    continue
                if should_translate_text(cell.value):
                    texts_to_translate.append(cell.value)

    unique_counter = Counter(normalize_text(text) for text in texts_to_translate if normalize_text(text))
    print(f"unique texts: {len(unique_counter)}", flush=True)
    translate_missing_texts(list(unique_counter.keys()), translator, cache, cache_path)

    translated_cells = 0
    skipped_cells = 0
    for ws in wb.worksheets:
        print(f"[sheet] {ws.title}", flush=True)
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    skipped_cells += 1
                    continue
                if not should_translate_text(cell.value):
                    skipped_cells += 1
                    continue
                original = cell.value
                translated = translate_text(original, translator, cache)
                if translated and translated != original:
                    cell.value = translated
                    translated_cells += 1
                    if translated_cells % 50 == 0:
                        print(f"  cells written: {translated_cells}", flush=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    save_json(cache_path, cache)
    print(f"translated cells: {translated_cells}", flush=True)
    print(f"skipped cells: {skipped_cells}", flush=True)
    print(f"saved: {output_path}", flush=True)


def main() -> None:
    ap = argparse.ArgumentParser(description="Translate English text in an Excel workbook into Chinese.")
    ap.add_argument("--input-xlsx", required=True, help="Input .xlsx or .xlsm file")
    ap.add_argument("--output-xlsx", required=True, help="Output .xlsx or .xlsm file")
    ap.add_argument(
        "--cache-json",
        default="tmp/excel_translate/cache_zh.json",
        help="Translation cache json path",
    )
    args = ap.parse_args()

    input_path = Path(args.input_xlsx).expanduser().resolve()
    output_path = Path(args.output_xlsx).expanduser().resolve()
    cache_path = Path(args.cache_json).expanduser().resolve()

    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise SystemExit("Only .xlsx and .xlsm are supported.")

    translate_workbook(input_path, output_path, cache_path)


if __name__ == "__main__":
    main()
