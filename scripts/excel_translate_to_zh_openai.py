#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import time
from collections import Counter
from pathlib import Path
from typing import Dict, List

import requests
from openpyxl import load_workbook


OPENAI_RESPONSES_URL = "https://api.openai.com/v1/responses"
XIAOMI_CHAT_URL = "https://api.xiaomimimo.com/v1/chat/completions"


def load_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def looks_like_non_natural_text(text: str) -> bool:
    s = text.strip()
    if not s:
        return True
    if re.fullmatch(r"https?://\S+", s, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"[\w.+-]+@[\w.-]+\.\w+", s):
        return True
    if re.fullmatch(r"[A-Z0-9_.\-/]{6,}", s):
        return True
    return False


def should_translate_text(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = normalize_text(value)
    if not text:
        return False
    if looks_like_non_natural_text(text):
        return False
    if not re.search(r"[A-Za-z]", text):
        return False
    return True


def build_workbook_context(input_path: Path) -> Dict[str, object]:
    wb = load_workbook(filename=str(input_path), read_only=True)
    sheet_info: List[Dict[str, object]] = []
    samples: List[str] = []
    headings: List[str] = []

    for ws in wb.worksheets:
        nonempty = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip():
                    nonempty += 1
                    text = normalize_text(value)
                    if len(samples) < 40:
                        samples.append(text)
                    if len(text) <= 80 and re.search(r"[A-Za-z]", text):
                        headings.append(text)
        sheet_info.append(
            {
                "name": ws.title,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "nonempty_text_cells": nonempty,
            }
        )

    unique_headings = list(dict.fromkeys(headings))[:60]
    return {
        "file_name": input_path.name,
        "sheet_info": sheet_info,
        "sample_texts": samples,
        "candidate_headings": unique_headings,
    }


def call_model(provider: str, model: str, instructions: str, input_text: str, max_output_tokens: int = 5000) -> str:
    if provider == "openai":
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY is not set")
        url = OPENAI_RESPONSES_URL
        payload = {
            "model": model,
            "instructions": instructions,
            "input": input_text,
            "max_output_tokens": max_output_tokens,
            "text": {"format": {"type": "text"}},
        }
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
    elif provider == "xiaomi":
        api_key = os.getenv("MIMO_API_KEY")
        if not api_key:
            raise RuntimeError("MIMO_API_KEY is not set")
        url = XIAOMI_CHAT_URL
        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": instructions},
                {"role": "user", "content": input_text},
            ],
            "temperature": 0.2,
            "max_tokens": max_output_tokens,
            "response_format": {"type": "json_object"},
        }
        headers = {
            "api-key": api_key,
            "Content-Type": "application/json",
        }
    else:
        raise RuntimeError(f"Unsupported provider: {provider}")

    last_error = None
    for _ in range(4):
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=180)
            if resp.status_code >= 400:
                raise RuntimeError(f"OpenAI API error {resp.status_code}: {resp.text[:800]}")
            data = resp.json()
            if provider == "openai":
                output_text = data.get("output_text")
                if output_text:
                    return output_text
            else:
                choices = data.get("choices") or []
                if choices:
                    message = choices[0].get("message") or {}
                    content = message.get("content")
                    if isinstance(content, str) and content.strip():
                        return content
            raise RuntimeError(f"Unexpected response shape: {json.dumps(data)[:1200]}")
        except Exception as exc:
            last_error = exc
            time.sleep(2.0)
    raise RuntimeError(str(last_error))


def build_translation_brief(context: Dict[str, object], model: str, provider: str) -> str:
    instructions = (
        "You are a professional EN->zh-CN translator. "
        "Read the workbook metadata and sample texts, infer the business domain, document purpose, tone, and recurring terminology. "
        "Return strict JSON only with keys: domain, purpose, tone, glossary, style_rules. "
        "glossary must be an array of objects with keys en and zh. "
        "style_rules must be an array of short Chinese strings. "
        "Preserve technical abbreviations when appropriate, and use concise Mainland Chinese suitable for supplier quality assessment forms."
    )
    input_text = json.dumps(context, ensure_ascii=False, indent=2)
    raw = call_model(provider=provider, model=model, instructions=instructions, input_text=input_text, max_output_tokens=2500)
    return raw


def strip_code_fence(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        lines = text.splitlines()
        if len(lines) >= 3:
            return "\n".join(lines[1:-1]).strip()
    return text


def parse_json_text(text: str):
    text = strip_code_fence(text)
    return json.loads(text)


def translate_batch(
    texts: List[str],
    brief: Dict[str, object],
    model: str,
    provider: str,
) -> Dict[str, str]:
    instructions = (
        "You are translating cells from one Excel workbook from English to zh-CN. "
        "All cells belong to the same document and must stay terminologically consistent. "
        "Follow the provided document brief and glossary. "
        "Return strict JSON only in the form {\"translations\":[{\"source\":\"...\",\"target\":\"...\"}, ...]}. "
        "Keep line breaks and bullet structures when they carry meaning. "
        "Do not explain anything. Do not omit any item. "
        "Preserve acronyms such as APQP, PPAP, FMEA, KPI, ERP, MRP, RoHS, TPM, RFID unless Chinese usage normally mixes them with Chinese."
    )
    input_payload = {
        "document_brief": brief,
        "texts": texts,
    }
    raw = call_model(
        provider=provider,
        model=model,
        instructions=instructions,
        input_text=json.dumps(input_payload, ensure_ascii=False, indent=2),
        max_output_tokens=5000,
    )
    data = parse_json_text(raw)
    pairs = data.get("translations", [])
    out: Dict[str, str] = {}
    for item in pairs:
        source = normalize_text(item.get("source", ""))
        target = normalize_text(item.get("target", ""))
        if source:
            out[source] = target or source
    missing = [text for text in texts if normalize_text(text) not in out]
    if missing:
        raise RuntimeError(f"Model omitted {len(missing)} items in batch")
    return out


def translate_workbook(input_path: Path, output_path: Path, cache_path: Path, model: str, provider: str) -> None:
    wb = load_workbook(
        filename=str(input_path),
        keep_vba=input_path.suffix.lower() == ".xlsm",
    )
    cache: Dict[str, str] = load_json(cache_path, {})

    texts_to_translate: List[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    continue
                if should_translate_text(cell.value):
                    texts_to_translate.append(cell.value)

    unique_texts = list(Counter(normalize_text(text) for text in texts_to_translate if normalize_text(text)).keys())
    print(f"unique texts: {len(unique_texts)}", flush=True)

    context = build_workbook_context(input_path)
    brief_path = cache_path.with_name("document_brief.json")
    if brief_path.exists():
        brief = load_json(brief_path, {})
    else:
        brief = parse_json_text(build_translation_brief(context, model, provider))
        save_json(brief_path, brief)
    print(f"document domain: {brief.get('domain', '')}", flush=True)

    missing = [text for text in unique_texts if text not in cache]
    batch_size = 20
    for idx in range(0, len(missing), batch_size):
        batch = missing[idx : idx + batch_size]
        translated = translate_batch(batch, brief, model, provider)
        cache.update(translated)
        save_json(cache_path, cache)
        print(f"  batch translated: {min(idx + len(batch), len(missing))}/{len(missing)}", flush=True)

    translated_cells = 0
    skipped_cells = 0
    for ws in wb.worksheets:
        print(f"[sheet] {ws.title}", flush=True)
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    skipped_cells += 1
                    continue
                if not should_translate_text(cell.value):
                    skipped_cells += 1
                    continue
                original = normalize_text(cell.value)
                translated = cache.get(original, original)
                if translated and translated != original:
                    cell.value = translated
                    translated_cells += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"translated cells: {translated_cells}", flush=True)
    print(f"skipped cells: {skipped_cells}", flush=True)
    print(f"saved: {output_path}", flush=True)


def main() -> None:
    ap = argparse.ArgumentParser(description="Translate an Excel workbook into Chinese using the OpenAI Responses API.")
    ap.add_argument("--input-xlsx", required=True)
    ap.add_argument("--output-xlsx", required=True)
    ap.add_argument("--model", default="gpt-5.2")
    ap.add_argument("--provider", choices=["openai", "xiaomi"], default="openai")
    ap.add_argument("--cache-json", default="tmp/excel_translate_openai/cache_zh.json")
    args = ap.parse_args()

    input_path = Path(args.input_xlsx).expanduser().resolve()
    output_path = Path(args.output_xlsx).expanduser().resolve()
    cache_path = Path(args.cache_json).expanduser().resolve()

    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise SystemExit("Only .xlsx and .xlsm are supported.")

    translate_workbook(input_path, output_path, cache_path, args.model, args.provider)


if __name__ == "__main__":
    main()
